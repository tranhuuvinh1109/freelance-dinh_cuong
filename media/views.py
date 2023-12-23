import mimetypes
from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
import xlsxwriter
import openpyxl
from django.core.exceptions import SuspiciousFileOperation
from openpyxl import load_workbook
import os
from django.core.files import File
from media.models import Media

from media.serializers import MediaSerializer
from . import action
from django.http import HttpResponse, HttpResponseNotFound
from django.utils.encoding import smart_str
from wsgiref.util import FileWrapper

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PATH_SAVE_DIR = os.path.join(BASE_DIR, 'manage/media')
class Check(APIView):
    def get(self, request):
        try:

            xlsx_files = [f for f in os.listdir(PATH_SAVE_DIR) if f.endswith('.xlsx') and os.path.isfile(os.path.join(PATH_SAVE_DIR, f))]

            return Response({'message': 'Running...', 'xlsx_files': xlsx_files}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CreateMediaSheet(APIView):
    def post(self, request):
        try:
            name = request.data.get('name')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')

            os.makedirs(os.path.dirname(path), exist_ok=True)

            workbook = xlsxwriter.Workbook(path)
            worksheet = workbook.add_worksheet()
            worksheet.write('A1', 'Hello..')
            worksheet.write('B1', 'Geeks')
            worksheet.write('C1', 'For')
            worksheet.write('D1', 'Geeks')

            workbook.close()

            return Response({'message': 'ok', 'name': name + '.xlsx'}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class ClearSheet(APIView):
    def post(self, request):
        try:
            name = request.data.get('name')
            sheet_name = request.data.get('sheet_name') 
            row_start = request.data.get('row_start')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')

            os.makedirs(os.path.dirname(path), exist_ok=True)

            res = action.clear_data(path, sheet_name, row_start)
            return Response(res, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetDataSheet(APIView):
    def get(self, request, *args, **kwargs):
        try:
            name = kwargs.get('name')
            sheet = kwargs.get('sheet')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')
            res = action.get_data_sheet(path, sheet)
            

            return Response(res, status=status.HTTP_200_OK)

        except SuspiciousFileOperation as e:
            return Response({'message': str(e)}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class InsertData(APIView):
    def post(self, request):
        try:
            column_A = request.data.get('column_A')
            column_B = request.data.get('column_B')
            column_C = request.data.get('column_C')
            column_D = request.data.get('column_D')
            name = request.data.get('name')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')

            workbook = load_workbook(path)
            worksheet = workbook.active

            new_row = [column_A, column_B, column_C, column_D]
            worksheet.append(new_row)
            workbook.save(path)

            return Response({'message': 'Data inserted successfully', "file": path}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class InsertDataLocation(APIView):
    def post(self, request):
        try:
            location = request.data.get('location')
            value = request.data.get('value')
            sheet_name = request.data.get('sheet') 

            # Get the file name from the request data (you may modify this part based on your needs)
            name = request.data.get('name')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')

            res = action.insert(path, sheet_name, location, value)
            return Response(res, status=status.HTTP_200_OK)

        except FileNotFoundError as e:
            return Response({'message': str(e)}, status=status.HTTP_404_NOT_FOUND)

        except ValueError as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class MergeCells(APIView):
    def post(self, request):
        try:
            start_cell = request.data.get('startCell')
            end_cell = request.data.get('endCell')
            name = request.data.get('name')
            sheet = request.data.get('sheetName')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')

            res = action.merge_cell(path,sheet, start_cell, end_cell )

            return Response(res, status=status.HTTP_200_OK)

        except FileNotFoundError as e:
            return Response({'message': str(e)}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class CreateNewSheet(APIView):
    def post(self, request):
        try:
            name = request.data.get('name')
            path = os.path.join(PATH_SAVE_DIR, name + '.xlsx')
            new_sheet_name = request.data.get('newSheetName')
            res = action.create_new_sheet(path, new_sheet_name)

            return Response(res, status=status.HTTP_200_OK)

        except FileNotFoundError as e:
            return Response({'message': str(e)}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# class Download(View):
#     def get(self, request):
#         file_name = 'media.xlsx'
#         file_path = os.path.join(BASE_DIR, 'manage/media', file_name)

#         # Fetch data from the Media model
#         media_instances = Media.objects.all()

#         # Serialize the data
#         serializer = MediaSerializer(media_instances, many=True)

#         # Convert the serialized data to a pandas DataFrame
#         data_frame = pd.DataFrame(serializer.data)

#         # Create a new Excel workbook
#         workbook = Workbook()
#         sheet = workbook.active

#         # Write the data to the Excel file
#         for idx, row in enumerate(data_frame.iterrows(), start=2):
#             for col_idx, value in enumerate(row[1], start=1):
#                 sheet.cell(row=idx, column=col_idx, value=value)

#         # Save the workbook to the file path
#         workbook.save(file_path)

#         try:
#             # Read the file data
#             with open(file_path, 'rb') as f:
#                 file_data = f.read()

#             # Sending response
#             response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
#             response['Content-Disposition'] = f'attachment; filename="{file_name}"'

#         except IOError:
#             # Handle file not exist case here
#             response = HttpResponseNotFound('<h1>File not exist</h1>')

#         return response
class CreateMedia(APIView):
    def post(self, request, *args, **kwargs):
        serializer = MediaSerializer(data=request.data)

        try:
            
            serializer.is_valid(raise_exception=True)

            serializer.save()
            file_name = 'media.xlsx'
            file_path = os.path.join(BASE_DIR, 'manage/media', file_name)
            values = [value for key, value in serializer.data.items() if key not in ['id', 'createAt', 'updateAt']]
            res = action.insert_multiple(file_path, 'media',values )

            return Response({'data': serializer.data, 'res': res}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class GetAllMedia(APIView):
    def get(self, request, *args, **kwargs):
        media_instances = Media.objects.all()

        serializer = MediaSerializer(media_instances, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)
class GetMediaByID(APIView):
    def get(self, request, pk, *args, **kwargs):
        try:
            media_instance = Media.objects.get(pk=pk)

            serializer = MediaSerializer(media_instance)

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Media.DoesNotExist:
            return Response({'message': 'Media not found'}, status=status.HTTP_404_NOT_FOUND)

class UpdateMedia(APIView):
    def put(self, request, pk, *args, **kwargs):
        try:
            media_instance = Media.objects.get(pk=pk)

            serializer = MediaSerializer(media_instance, data=request.data)

            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Media.DoesNotExist:
            return Response({'message': 'Media not found'}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)
class DeleteAllMedia(APIView):
    def delete(self, request, *args, **kwargs):
        try:
            # Delete all instances of Media
            Media.objects.all().delete()

            return Response({'message': 'All media deleted successfully'}, status=status.HTTP_204_NO_CONTENT)

        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# class CheckWorker(APIView):
#     def get(self, request):
#         return Response({'message': 'Running...'}, status=status.HTTP_200_OK)

    
