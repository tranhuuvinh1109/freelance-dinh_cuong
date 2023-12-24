import mimetypes
from openpyxl import load_workbook
import os
from django.core.exceptions import SuspiciousFileOperation
from django.http import FileResponse, HttpResponse
from openpyxl import Workbook
import os
from wsgiref.util import FileWrapper
from django.utils.encoding import smart_str

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MEDIA_DIR = os.path.join(BASE_DIR, 'manage/media')

def get_data_sheet(file, sheet):
    try:
        if not os.path.exists(file):
            raise SuspiciousFileOperation("File not found")

        workbook = load_workbook(file)

        if sheet not in workbook.sheetnames:
            return {'message': f"Sheet '{sheet}' not found in the workbook"}

        worksheet = workbook[sheet]

        data = []
        for row in worksheet.iter_rows(min_row=6, values_only=True):
            data.append(row)

        return {'message': 'Get data sheet success', 'file': file, 'sheet': sheet, 'data': data}

    except SuspiciousFileOperation as e:
        return {'message': str(e)}

    except Exception as e:
        return {'message': f"An error occurred: {str(e)}"}

def insert(file, sheet, location, value):
    if not os.path.exists(file):
        return FileNotFoundError("File not found")

    try:
        workbook = load_workbook(file)
    except Exception as e:
        return {'message': f"Error loading workbook: {str(e)}"}

    if sheet not in workbook.sheetnames:
        return {'message': f"Sheet '{sheet}' not found in the workbook"}

    worksheet = workbook[sheet]
    worksheet[location] = value
    workbook.save(file)
    workbook.close()

    return {'message': 'Insert success', 'sheet_name': sheet, 'updated_location': location, 'updated_value': value}

def insert_multiple(file, sheet, rows_of_values):
    if not os.path.exists(file):
        return FileNotFoundError("File not found")

    try:
        workbook = load_workbook(file)
    except Exception as e:
        return {'message': f"Error loading workbook: {str(e)}"}

    if sheet not in workbook.sheetnames:
        return {'message': f"Sheet '{sheet}' not found in the workbook"}

    worksheet = workbook[sheet]

    last_row = worksheet.max_row + 1

    for col_num, value in enumerate(rows_of_values, start=1):
        cell_location = f"{chr(ord('A') + col_num - 1)}{last_row}"
        worksheet[cell_location] = value

    workbook.save(file)
    workbook.close()

    return {'message': 'Insert success', 'sheet_name': sheet, 'values': rows_of_values}



def has_data_from_row(file, sheet, row_number):
    try:
        workbook = load_workbook(file)
        worksheet = workbook[sheet]

        # Check if any cell in the specified row has data
        for col_num in range(1, worksheet.max_column + 1):
            cell_location = f"{chr(ord('A') + col_num - 1)}{row_number}"
            if worksheet[cell_location].value is not None:
                workbook.close()
                return True
        workbook.close()

        return False

    except Exception as e:
        # Handle exceptions as needed
        print(f"Error checking data: {str(e)}")
        return False
    

def create_new_sheet(file, new_sheet_name):
    try:
        if not os.path.exists(file):
            raise FileNotFoundError("File not found")

        workbook = load_workbook(file)
        new_sheet = workbook.create_sheet(title=new_sheet_name)
        workbook.save(file)
        workbook.close()

        return {'message': 'Create new sheet success', 'file_name': file ,'new_sheet_name': new_sheet_name}

    except FileNotFoundError as e:
        return {'message': str(e)}

    except Exception as e:
        return {'message': f"An error occurred: {str(e)}"}
    
def merge_cell(file, sheet, start_cell, end_cell):
    try:
        if not os.path.exists(file):
            raise FileNotFoundError("File not found")

        workbook = load_workbook(file)

        if sheet not in workbook.sheetnames:
            return {'message': f"Sheet '{sheet}' not found in the workbook"}

        worksheet = workbook[sheet]

        worksheet.merge_cells(f"{start_cell}:{end_cell}")

        workbook.save(file)
        workbook.close()

        return {'message': 'Merge cell success', 'merged_cells': f"{start_cell}:{end_cell}"}

    except FileNotFoundError as e:
        return {'message': str(e)}

    except Exception as e:
        return {'message': f"An error occurred: {str(e)}"}
    
    
    
# def download_excel(path):
#     try:
#         file_path = os.path.join('your-path-to-excel-files', path)

#         if not os.path.exists(file_path):
#             raise SuspiciousFileOperation("File not found")

#         excel_file = open(file_path, 'rb')
#         response = FileResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = f'attachment; filename="{path}"'

#         return response

#     except SuspiciousFileOperation as e:
#         return (str(e))
    
def download_excel(file_name='media.xlsx'):
    file_path = os.path.join(BASE_DIR, 'manage/media', file_name)

    with open(file_path, 'rb') as file:
        file_wrapper = FileWrapper(file)
        file_mimetype, _ = mimetypes.guess_type(file_path)

        response = HttpResponse(file_wrapper, content_type=file_mimetype)
        response['X-Sendfile'] = file_path
        response['Content-Length'] = os.path.getsize(file_path)
        response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(file_name)

    return response

def clear_data(file, sheet, start_row):
    try:
        if not os.path.exists(file):
            raise FileNotFoundError("File not found")

        workbook = load_workbook(file)

        if sheet not in workbook.sheetnames:
            return {'message': f"Sheet '{sheet}' not found in the workbook"}

        worksheet = workbook[sheet]
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = None

        workbook.save(file)
        workbook.close()

        return {'message': f'Data cleared from row {start_row} onwards in {sheet}'}

    except FileNotFoundError as e:
        return {'message': str(e)}

    except Exception as e:
        return {'message': f"An error occurred: {str(e)}"}
    
def insert_data_into_excel(file, sheet, data):
    try:
        # If the file does not exist, create a new workbook
        if not os.path.exists(file):
            workbook = Workbook()
            workbook.save(file)

        workbook = load_workbook(file)

    except Exception as e:
        return {'message': f"Error loading workbook: {str(e)}"}

    if sheet not in workbook.sheetnames:
        workbook.create_sheet(title=sheet)

    worksheet = workbook[sheet]

    # If the sheet is empty, write headers
    if worksheet.max_row == 1:
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, start=1):
            cell_location = f"{chr(ord('A') + col_num - 1)}1"
            worksheet[cell_location] = header

    # Get the last row in the sheet
    last_row = worksheet.max_row + 1

    # Iterate through the data and insert into the sheet
    for row_data in data:
        for col_num, (header, value) in enumerate(row_data.items(), start=1):
            cell_location = f"{chr(ord('A') + col_num - 1)}{last_row}"
            worksheet[cell_location] = value

        # Move to the next row
        last_row += 1

    # Save the changes to the workbook
    workbook.save(file)
    workbook.close()

    return {'message': 'Insert success', 'sheet_name': sheet, 'values': data}

def clear_data_sheet(file, sheet, start_row=2):
    try:
        if not os.path.exists(file):
            return {'message': 'File not found'}

        workbook = load_workbook(file)

    except Exception as e:
        return {'message': f"Error loading workbook: {str(e)}"}

    if sheet not in workbook.sheetnames:
        return {'message': f"Sheet '{sheet}' not found in the workbook"}

    worksheet = workbook[sheet]

    for row in worksheet.iter_rows(min_row=start_row):
        for cell in row:
            cell.value = None

    workbook.save(file)
    workbook.close()

    return {'message': f'Cleared data from row {start_row} onward', 'sheet_name': sheet}
